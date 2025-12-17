from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
from .models import RegistroRacion, RegistroUsuario, ProyeccionRacion
from .models import Insumo  # Asegúrate de ponerlo arriba con tus otros imports
from .models import ProduccionHuevos
import json
from decimal import Decimal, InvalidOperation
from datetime import date
from collections import defaultdict

from django.contrib.auth.decorators import login_required 
from django.shortcuts import render, redirect 
from .models import Insumo


from decimal import Decimal 
from django.shortcuts import get_object_or_404, render, redirect 
from django.contrib.auth.decorators import login_required 
from .models import Insumo
# ================== LOGIN ==================
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            return render(request, "panel/login.html", {"error": "Usuario o contraseña incorrecta"})

    return render(request, "panel/login.html")


# ================== REGISTRO ==================
def register_view(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre_completo")
        username = request.POST.get("username")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            return render(request, "panel/register.html", {"error": "Las contraseñas no coinciden"})

        if User.objects.filter(username=username).exists():
            return render(request, "panel/register.html", {"error": "Usuario ya registrado"})

        User.objects.create_user(
            username=username,
            password=password1,
            first_name=nombre
        )
        return redirect("login")

    return render(request, "panel/register.html")


# ================== LOGOUT ==================
def logout_view(request):
    logout(request)
    return redirect("login")


# ================== PÁGINAS PROTEGIDAS ==================
@login_required
@login_required
def dashboard(request):
    # Datos para gráfico dinámico de producción de huevos
    producciones_for_chart = ProduccionHuevos.objects.all().order_by('fecha')
    data_by_date = defaultdict(lambda: {'PEQUEÑO': 0, 'MEDIANO': 0, 'GRANDE': 0, 'EXTRA_GRANDE': 0})
    for prod in producciones_for_chart:
        data_by_date[prod.fecha][prod.tipo_huevo] += prod.cantidad
    fechas = sorted([fecha.strftime('%Y-%m-%d') for fecha in data_by_date.keys()])
    tipos = {}
    for tipo in ['PEQUEÑO', 'MEDIANO', 'GRANDE', 'EXTRA_GRANDE']:
        tipos[tipo] = [data_by_date[date.fromisoformat(fecha)][tipo] for fecha in fechas]

    return render(request, "panel/dashboard.html", {
        "fechas_json": json.dumps(fechas),
        "tipos_json": json.dumps(tipos)
    })

@login_required
def raciones(request):
    dias = [
        ("Lunes", "L"), ("Martes", "M"), ("Miércoles", "X"),
        ("Jueves", "J"), ("Viernes", "V"), ("Sábado", "S"), ("Domingo", "D")
    ]
    return render(request, "panel/raciones.html", {"dias": dias})

@login_required
def proyecciones(request):
    # Obtener todas las raciones guardadas
    raciones = RegistroRacion.objects.all().order_by('-creado_en')

    # Crear opciones solo con tipo de animal y peso
    opciones = []
    raciones_data = {}
    for r in raciones:
        opciones.append({
            "id": r.id,
            "descripcion": f"{r.tipo_animal} - {r.peso} kg"
        })
        # incluir datos relevantes por ración para uso en JS
        raciones_data[r.id] = {
            "granos": float(r.granos) if r.granos is not None else 0,
            "algas": float(r.algas) if r.algas is not None else 0,
            "peso": float(r.peso) if r.peso is not None else 0,
            "dias": r.dias,
        }

    # Datos para gráfico dinámico de producción de huevos
    producciones = ProduccionHuevos.objects.all().order_by('fecha')
    fechas = []
    cantidades = []
    tipos = {'PEQUEÑO': [], 'MEDIANO': [], 'GRANDE': [], 'EXTRA_GRANDE': []}
    for prod in producciones:
        fechas.append(prod.fecha.strftime('%Y-%m-%d'))
        cantidades.append(prod.cantidad)
        for tipo in tipos:
            tipos[tipo].append(prod.cantidad if prod.tipo_huevo == tipo else 0)

    return render(request, "panel/proyecciones.html", {
        "raciones": opciones,
        "raciones_data_json": json.dumps(raciones_data),
        "fechas_json": json.dumps(fechas),
        "cantidades_json": json.dumps(cantidades),
        "tipos_json": json.dumps(tipos)
    })

@login_required
def registros(request):
    registros = RegistroRacion.objects.all()
    return render(request, "panel/registros.html", {"registros": registros})

@login_required
def configuraciones(request):
    return render(request, "panel/configuraciones.html")

@login_required
def perfil_usuario(request):
    return render(request, "panel/perfil_usuario.html")


# ================== REGISTROS ==================
@login_required
def registro_eliminar(request, id):
    registro = get_object_or_404(RegistroRacion, id=id)
    registro.delete()
    return redirect("registros")

@login_required
def registro_editar(request, id):
    registro = get_object_or_404(RegistroRacion, id=id)

    if request.method == "POST":
        tipo_animal = request.POST.get("tipo_animal")
        peso_str = request.POST.get("peso")
        granos_str = request.POST.get("granos")
        algas_str = request.POST.get("algas")
        dias = request.POST.get("dias")

        # Validar campos decimales
        try:
            peso = Decimal(peso_str)
            if peso < 0 or peso >= 10000:  # max_digits=6, decimal_places=2 -> max 9999.99
                raise ValueError("Valor fuera de rango")
        except (InvalidOperation, ValueError):
            return render(request, "panel/editar_registro.html", {
                "registro": registro,
                "dias": [("Lunes", "L"), ("Martes", "M"), ("Miércoles", "X"), ("Jueves", "J"), ("Viernes", "V"), ("Sábado", "S"), ("Domingo", "D")],
                "error": "El peso debe ser un número válido entre 0 y 9999.99"
            })

        try:
            granos = Decimal(granos_str)
            if granos < 0 or granos >= 10000:
                raise ValueError("Valor fuera de rango")
        except (InvalidOperation, ValueError):
            return render(request, "panel/editar_registro.html", {
                "registro": registro,
                "dias": [("Lunes", "L"), ("Martes", "M"), ("Miércoles", "X"), ("Jueves", "J"), ("Viernes", "V"), ("Sábado", "S"), ("Domingo", "D")],
                "error": "Los granos deben ser un número válido entre 0 y 9999.99"
            })

        try:
            algas = Decimal(algas_str)
            if algas < 0 or algas >= 10000:
                raise ValueError("Valor fuera de rango")
        except (InvalidOperation, ValueError):
            return render(request, "panel/editar_registro.html", {
                "registro": registro,
                "dias": [("Lunes", "L"), ("Martes", "M"), ("Miércoles", "X"), ("Jueves", "J"), ("Viernes", "V"), ("Sábado", "S"), ("Domingo", "D")],
                "error": "Las algas deben ser un número válido entre 0 y 9999.99"
            })

        registro.tipo_animal = tipo_animal
        registro.peso = peso
        registro.granos = granos
        registro.algas = algas
        registro.dias = dias
        registro.save()
        return redirect("registros")

    dias = [
        ("Lunes", "L"), ("Martes", "M"), ("Miércoles", "X"),
        ("Jueves", "J"), ("Viernes", "V"), ("Sábado", "S"), ("Domingo", "D")
    ]
    return render(request, "panel/editar_registro.html", {"registro": registro, "dias": dias})


# ================== GUARDAR RACIONES ==================
@login_required
def guardar_racion(request):
    if request.method == "POST":
        try:
            print("DEBUG: Iniciando guardar_racion")
            print(f"DEBUG: Headers: {dict(request.headers)}")
            print(f"DEBUG: POST data: {dict(request.POST)}")
            
            tipo = request.POST.get("tipo_animal")
            peso_str = request.POST.get("peso")
            granos_str = request.POST.get("granos")
            algas_str = request.POST.get("algas")
            dias_list = request.POST.getlist("dias")  # lista de días seleccionados

            print(f"DEBUG: tipo={tipo}, peso_str={peso_str}, granos_str={granos_str}, algas_str={algas_str}, dias_list={dias_list}")

            errors = []

            # Validar tipo_animal
            if not tipo or len(tipo) > 50:
                errors.append("Tipo de animal inválido")

            # Validar dias
            if not dias_list:
                errors.append("Debe seleccionar al menos un día")

            dias = " - ".join(dias_list)
            if len(dias) > 60:
                errors.append("Demasiados días seleccionados")

            # Validar campos decimales
            try:
                peso = Decimal(peso_str)
                if peso <= 0 or peso >= 10000:
                    errors.append("Peso debe estar entre 0.01 y 9999.99")
            except (InvalidOperation, ValueError):
                errors.append("Peso debe ser un número válido")

            try:
                granos = Decimal(granos_str)
                if granos < 0 or granos >= 10000:
                    errors.append("Granos debe estar entre 0 y 9999.99")
            except (InvalidOperation, ValueError):
                errors.append("Granos debe ser un número válido")

            try:
                algas = Decimal(algas_str)
                if algas < 0 or algas >= 10000:
                    errors.append("Algas debe estar entre 0 y 9999.99")
            except (InvalidOperation, ValueError):
                errors.append("Algas debe ser un número válido")

            print(f"DEBUG: errors={errors}")

            # Si hay errores, retornar JSON para AJAX
            if errors:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    print("DEBUG: Retornando errores JSON")
                    return JsonResponse({'success': False, 'errors': errors})
                return redirect("raciones")

            registro = RegistroRacion.objects.create(
                tipo_animal=tipo,
                peso=peso,
                granos=granos,
                algas=algas,
                dias=dias
            )

            # Si es petición AJAX, devuelve JSON de éxito
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                print("DEBUG: Retornando éxito JSON")
                return JsonResponse({"success": True, "id": registro.id})

            return redirect("registros")

        except Exception as e:
            print(f"DEBUG: Exception caught: {e}")
            import traceback
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            # Capturar cualquier excepción y retornar JSON para AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                print("DEBUG: Retornando error JSON por excepción")
                return JsonResponse({'success': False, 'errors': ['Error interno del servidor']})
            return redirect("raciones")

    return redirect("raciones")


from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Insumo



def stock(request):
    insumos = Insumo.objects.all()
    alertas_stock = []

    for insumo in insumos:
        minimo = insumo.stock_minimo
        actual = insumo.stock_actual

        if actual <= minimo * Decimal('0.5'):
            nivel = 'Crítico'
        elif actual <= minimo * Decimal('1.5'):
            nivel = 'Bajo'
        else:
            nivel = 'Normal'

        # 🔴 agrega el estado al objeto insumo
        insumo.estado = nivel

        alertas_stock.append({
            'nombre': insumo.nombre,
            'stock_actual': actual,
            'nivel': nivel
        })

    return render(request, "panel/stock.html", {
        'insumos': insumos,
        'alertas_stock': alertas_stock
    })


from decimal import Decimal
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from .models import Insumo

@login_required
def stock_agregar(request, id):
    insumo = get_object_or_404(Insumo, id=id)
    if request.method == "POST":
        try:
            # Convertir a Decimal en vez de float
            cantidad = Decimal(request.POST.get("cantidad", "0").strip())

            if cantidad < 0:
                return render(request, "panel/stock_agregar.html", {
                    "insumo": insumo,
                    "error": "La cantidad no puede ser negativa"
                })

            insumo.stock_actual += cantidad   # ✅ Decimal + Decimal
            insumo.save()
            return redirect("stock")

        except (ValueError, ArithmeticError):
            return render(request, "panel/stock_agregar.html", {
                "insumo": insumo,
                "error": "Ingresa un número válido"
            })

    return render(request, "panel/stock_agregar.html", {"insumo": insumo})

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Insumo

@login_required
def stock_crear(request):
    """Crear un nuevo insumo"""
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        unidad = request.POST.get("unidad", "kg").strip()
        stock_actual = float(request.POST.get("stock_actual", 0))
        stock_minimo = float(request.POST.get("stock_minimo", 0))
        porcentaje_bodega = float(request.POST.get("porcentaje_bodega", 0))

        if not nombre:
            return render(request, "panel/stock_crear.html", {
                "error": "El nombre del insumo es requerido"
            })

        if Insumo.objects.filter(nombre=nombre).exists():
            return render(request, "panel/stock_crear.html", {
                "error": f"El insumo '{nombre}' ya existe"
            })

        Insumo.objects.create(
            nombre=nombre,
            unidad=unidad,
            stock_actual=stock_actual,
            stock_minimo=stock_minimo,
            porcentaje_ocupado_bodega=porcentaje_bodega
        )
        return redirect("stock")  # redirige a la vista de listado

    # 🔴 Aquí estaba faltando: devolver el formulario vacío en GET
    return render(request, "panel/stock_crear.html")



# ================== PRODUCCIÓN DE HUEVOS ==================
@login_required
def produccion_huevos(request):
    if request.method == "POST":
        fecha_str = request.POST.get("fecha")
        cantidad_str = request.POST.get("cantidad")
        tipo_huevo = request.POST.get("tipo_huevo") or 'MEDIANO'

        # Fecha es automática, no validar
        fecha = date.today()

        # Validar cantidad
        try:
            cantidad = int(cantidad_str)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            producciones = ProduccionHuevos.objects.all().order_by('-fecha')
            return render(request, "panel/produccion_huevos.html", {
                "producciones": producciones,
                "today": date.today(),
                "error": "La cantidad debe ser un número entero positivo"
            })

        # Validar tipo_huevo
        if tipo_huevo not in ['PEQUEÑO', 'MEDIANO', 'GRANDE', 'EXTRA_GRANDE']:
            producciones = ProduccionHuevos.objects.all().order_by('-fecha')
            return render(request, "panel/produccion_huevos.html", {
                "producciones": producciones,
                "today": date.today(),
                "error": "Tipo de huevo inválido"
            })

        # Verificar si ya existe un registro para esta fecha y tipo de huevo
        produccion_existente = ProduccionHuevos.objects.filter(fecha=fecha, tipo_huevo=tipo_huevo).first()
        
        if produccion_existente:
            # Si existe, sumar la cantidad
            produccion_existente.cantidad += cantidad
            produccion_existente.save()
        else:
            # Si no existe, crear nuevo registro
            ProduccionHuevos.objects.create(
                fecha=fecha,
                cantidad=cantidad,
                tipo_huevo=tipo_huevo
            )
        return redirect("produccion_huevos")

    producciones = ProduccionHuevos.objects.all().order_by('-fecha')

    # Datos para gráfico dinámico de producción de huevos
    producciones_for_chart = ProduccionHuevos.objects.all().order_by('fecha')
    data_by_date = defaultdict(lambda: {'PEQUEÑO': 0, 'MEDIANO': 0, 'GRANDE': 0, 'EXTRA_GRANDE': 0})
    for prod in producciones_for_chart:
        data_by_date[prod.fecha][prod.tipo_huevo] += prod.cantidad
    fechas = sorted([fecha.strftime('%Y-%m-%d') for fecha in data_by_date.keys()])
    tipos = {}
    for tipo in ['PEQUEÑO', 'MEDIANO', 'GRANDE', 'EXTRA_GRANDE']:
        tipos[tipo] = [data_by_date[date.fromisoformat(fecha)][tipo] for fecha in fechas]

    return render(request, "panel/produccion_huevos.html", {
        "producciones": producciones,
        "today": date.today(),
        "fechas_json": json.dumps(fechas),
        "tipos_json": json.dumps(tipos)
    })


@login_required
def produccion_huevos_agregar(request):
    if request.method == "POST":
        lote_id = request.POST.get("lote")
        fecha_str = request.POST.get("fecha")
        cantidad_str = request.POST.get("cantidad")
        tipo_huevo = request.POST.get("tipo_huevo") or 'MEDIANO'  # default

        try:
            cantidad = int(cantidad_str)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            return render(request, "panel/produccion_huevos_agregar.html", {
                "lotes": LoteGallinas.objects.filter(activo=True),
                "error": "La cantidad debe ser un número entero positivo"
            })

        # Convertir fecha
        try:
            fecha = date.fromisoformat(fecha_str)
        except ValueError:
            return render(request, "panel/produccion_huevos_agregar.html", {
                "lotes": LoteGallinas.objects.filter(activo=True),
                "error": "Fecha inválida"
            })

        lote = None
        if lote_id:
            lote = get_object_or_404(LoteGallinas, id=lote_id)

        # Verificar si ya existe un registro para esta fecha y tipo de huevo
        produccion_existente = ProduccionHuevos.objects.filter(fecha=fecha, tipo_huevo=tipo_huevo).first()
        
        if produccion_existente:
            # Si existe, sumar la cantidad
            produccion_existente.cantidad += cantidad
            produccion_existente.save()
        else:
            # Si no existe, crear nuevo registro
            ProduccionHuevos.objects.create(
                lote=lote,
                fecha=fecha,
                cantidad=cantidad,
                tipo_huevo=tipo_huevo
            )
        return redirect("produccion_huevos")

    lotes = LoteGallinas.objects.filter(activo=True)
    return render(request, "panel/produccion_huevos_agregar.html", {"lotes": lotes})


@login_required
def produccion_huevos_editar(request, id):
    produccion = get_object_or_404(ProduccionHuevos, id=id)

    if request.method == "POST":
        cantidad_str = request.POST.get("cantidad")
        tipo_huevo = request.POST.get("tipo_huevo")

        # Fecha es automática, no validar
        fecha = date.today()

        # Validar cantidad
        try:
            cantidad = int(cantidad_str)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            return render(request, "panel/produccion_huevos_editar.html", {
                "produccion": produccion,
                "error": "La cantidad debe ser un número entero positivo"
            })

        # Validar tipo_huevo
        if tipo_huevo not in ['PEQUEÑO', 'MEDIANO', 'GRANDE', 'EXTRA_GRANDE']:
            return render(request, "panel/produccion_huevos_editar.html", {
                "produccion": produccion,
                "error": "Tipo de huevo inválido"
            })

        # Si el tipo de huevo cambió, verificar si ya existe un registro para la nueva combinación
        if produccion.tipo_huevo != tipo_huevo or produccion.fecha != fecha:
            # Buscar si ya existe un registro para esta fecha y nuevo tipo
            produccion_existente = ProduccionHuevos.objects.filter(fecha=fecha, tipo_huevo=tipo_huevo).exclude(id=id).first()
            
            if produccion_existente:
                # Si existe, sumar la cantidad al registro existente y eliminar el actual
                produccion_existente.cantidad += cantidad
                produccion_existente.save()
                produccion.delete()
            else:
                # Si no existe, actualizar el registro actual
                produccion.fecha = fecha
                produccion.cantidad = cantidad
                produccion.tipo_huevo = tipo_huevo
                produccion.save()
        else:
            # Si no cambió el tipo ni la fecha, solo actualizar cantidad
            produccion.cantidad = cantidad
            produccion.save()
            
        return redirect("produccion_huevos")

    return render(request, "panel/produccion_huevos_editar.html", {"produccion": produccion, "today": date.today()})


@login_required
def produccion_huevos_eliminar(request, id):
    produccion = get_object_or_404(ProduccionHuevos, id=id)
    produccion.delete()
    return redirect("produccion_huevos")

    return render(request, "panel/stock_crear.html")


@login_required
def stock_editar(request, id):
    insumo = get_object_or_404(Insumo, id=id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        try:
            stock_actual = float(request.POST.get('stock_actual', insumo.stock_actual))
            stock_minimo = float(request.POST.get('stock_minimo', insumo.stock_minimo))
        except ValueError:
            return render(request, 'panel/editar_stock.html', {
                'insumo': insumo,
                'error': 'Valores numéricos inválidos'
            })

        if nombre:
            insumo.nombre = nombre
        insumo.stock_actual = stock_actual
        insumo.stock_minimo = stock_minimo
        insumo.save()
        return redirect('stock')

    return render(request, 'panel/editar_stock.html', {'insumo': insumo})


@login_required
def api_proyeccion_detalle(request, id):
    proyeccion = get_object_or_404(ProyeccionRacion, id=id)
    data = {
        'id': proyeccion.id,
        'racion_base': {
            'tipo_animal': proyeccion.racion_base.tipo_animal,
            'peso': float(proyeccion.racion_base.peso),
            'granos': float(proyeccion.racion_base.granos),
            'algas': float(proyeccion.racion_base.algas),
            'dias': proyeccion.racion_base.dias,
        },
        'cantidad_animales': proyeccion.cantidad_animales,
        'periodo_dias': proyeccion.periodo_dias,
        'unidad_racion': proyeccion.unidad_racion,
        'cantidades_maximas': {
            'granos_kg': float(proyeccion.total_granos_kg),
            'algas_kg': float(proyeccion.total_algas_kg),
            'carbonato_kg': float(proyeccion.total_carbonato_kg),
            'total_kg': float(proyeccion.total_granos_kg + proyeccion.total_algas_kg + proyeccion.total_carbonato_kg),
        },
        'cantidades_con_ahorro': {
            'granos_kg': float(proyeccion.total_granos_ahorro_kg),
            'algas_kg': float(proyeccion.total_algas_ahorro_kg),
            'carbonato_kg': float(proyeccion.total_carbonato_ahorro_kg),
            'total_kg': float(proyeccion.total_granos_ahorro_kg + proyeccion.total_algas_ahorro_kg + proyeccion.total_carbonato_ahorro_kg),
        },
        'ahorros': {
            'granos_kg': float(proyeccion.ahorro_granos_kg),
            'algas_kg': float(proyeccion.ahorro_algas_kg),
            'carbonato_kg': float(proyeccion.ahorro_carbonato_kg),
            'total_kg': float(proyeccion.ahorro_total_kg),
        },
        'creado_en': proyeccion.creado_en.isoformat(),
    }
    return JsonResponse(data)


@login_required
def lista_proyecciones(request):
    proyecciones = ProyeccionRacion.objects.all().order_by('-creado_en')
    return render(request, "panel/lista_proyecciones.html", {"proyecciones": proyecciones})


@login_required
def exportar_proyecciones_excel(request):
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyecciones de Consumo"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

    # Headers
    headers = [
        'Fecha de Creación', 'Ración Base', 'Animales', 'Periodo (días)', 'Unidad',
        'Granos Máximo (kg)', 'Algas Máximo (kg)', 'Carbonato Máximo (kg)', 'Total Máximo (kg)',
        'Granos con Ahorro (kg)', 'Algas con Ahorro (kg)', 'Carbonato con Ahorro (kg)', 'Total con Ahorro (kg)',
        'Ahorro Total (kg)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Datos
    proyecciones = ProyeccionRacion.objects.all().order_by('-creado_en')
    
    for row_num, proyeccion in enumerate(proyecciones, 2):
        ws.cell(row=row_num, column=1, value=proyeccion.creado_en.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=2, value=f"{proyeccion.racion_base.tipo_animal} - {proyeccion.racion_base.peso}kg")
        ws.cell(row=row_num, column=3, value=proyeccion.cantidad_animales)
        ws.cell(row=row_num, column=4, value=proyeccion.periodo_dias)
        ws.cell(row=row_num, column=5, value=proyeccion.unidad_racion)
        ws.cell(row=row_num, column=6, value=float(proyeccion.total_granos_kg))
        ws.cell(row=row_num, column=7, value=float(proyeccion.total_algas_kg))
        ws.cell(row=row_num, column=8, value=float(proyeccion.total_carbonato_kg))
        ws.cell(row=row_num, column=9, value=float(proyeccion.total_granos_kg + proyeccion.total_algas_kg + proyeccion.total_carbonato_kg))
        ws.cell(row=row_num, column=10, value=float(proyeccion.total_granos_ahorro_kg))
        ws.cell(row=row_num, column=11, value=float(proyeccion.total_algas_ahorro_kg))
        ws.cell(row=row_num, column=12, value=float(proyeccion.total_carbonato_ahorro_kg))
        ws.cell(row=row_num, column=13, value=float(proyeccion.total_granos_ahorro_kg + proyeccion.total_algas_ahorro_kg + proyeccion.total_carbonato_ahorro_kg))
        ws.cell(row=row_num, column=14, value=float(proyeccion.ahorro_total_kg))

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyecciones_consumo.xlsx'
    
    wb.save(response)
    return response


@login_required
def guardar_proyeccion(request):
    if request.method == "POST":
        try:
            print("DEBUG: Iniciando guardar_proyeccion")
            print(f"DEBUG: POST data: {dict(request.POST)}")
            
            racion_id = request.POST.get("racion_id")
            cantidad_animales = int(request.POST.get("cantidad_animales"))
            periodo_dias = int(request.POST.get("periodo_dias"))
            unidad_racion = request.POST.get("unidad_racion")
            
            # Cantidades máximas
            total_granos = Decimal(request.POST.get("total_granos"))
            total_algas = Decimal(request.POST.get("total_algas"))
            total_carbonato = Decimal(request.POST.get("total_carbonato"))
            
            # Cantidades con ahorro
            total_granos_ahorro = Decimal(request.POST.get("total_granos_ahorro"))
            total_algas_ahorro = Decimal(request.POST.get("total_algas_ahorro"))
            total_carbonato_ahorro = Decimal(request.POST.get("total_carbonato_ahorro"))
            ahorro_total = Decimal(request.POST.get("ahorro_total"))
            
            # Obtener la ración base
            racion = get_object_or_404(RegistroRacion, id=racion_id)
            
            # Crear la proyección
            proyeccion = ProyeccionRacion.objects.create(
                racion_base=racion,
                cantidad_animales=cantidad_animales,
                periodo_dias=periodo_dias,
                unidad_racion=unidad_racion,
                total_granos_kg=total_granos,
                total_algas_kg=total_algas,
                total_carbonato_kg=total_carbonato,
                total_granos_ahorro_kg=total_granos_ahorro,
                total_algas_ahorro_kg=total_algas_ahorro,
                total_carbonato_ahorro_kg=total_carbonato_ahorro,
                ahorro_granos_kg=total_granos - total_granos_ahorro,
                ahorro_algas_kg=total_algas - total_algas_ahorro,
                ahorro_carbonato_kg=total_carbonato - total_carbonato_ahorro,
                ahorro_total_kg=ahorro_total,
                creado_por=request.user
            )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({"success": True, "id": proyeccion.id})
            
            return redirect("proyecciones")
            
        except Exception as e:
            print(f"DEBUG: Exception in guardar_proyeccion: {e}")
            import traceback
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': ['Error interno del servidor']})
            return redirect("proyecciones")
    
    return redirect("proyecciones")